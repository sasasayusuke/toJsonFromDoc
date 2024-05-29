import util
import gui
import config
import openpyxl
import os
import re
import requests
import copy


def main(app):

    try:
        input_dir = app.entry_input_dir.get()
        output_dir = app.entry_output_dir.get()

        # ファイル名が '~$' で始まるものを除外し、正しい形式のものだけを処理
        interfaces = []
        pattern = r'([^_]+)_(.*)_画面設計書\.xlsx'
        for file_name in os.listdir(input_dir):
            if not file_name.startswith('~$') and re.match(pattern, file_name):
                matched = re.match(pattern, file_name)
                if matched:
                    # パターンにマッチした場合
                    site_id = matched.group(1)
                    interface_name = matched.group(2)
                    config.template_json["HeaderInfo"]["Convertors"][0]["SiteTitle"] = interface_name
                    interfaces.append({
                        "site_id": site_id,
                        "interface_name": interface_name,
                        "file_name": file_name
                    })

        app.log_output(f"# 読み取り件数: {len(interfaces)}")
        for interface in interfaces:
            app.log_output(f"## {interface['file_name']}")

        for interface in interfaces:
            file_path = os.path.join(input_dir, interface["file_name"])
            # ワークブックを読み込みを開始
            app.log_output(f"### {interface['file_name']} の Book読み込みを開始")
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # スケルトンのコピーを作成
            if app.radio_value.get() == "JSON":
                site_json = copy.deepcopy(config.site_skeleton_json)
            elif app.radio_value.get() == "API":
                site_json = copy.deepcopy(config.site_post_json)

                api_key = app.entry_api_key.get()
                server = app.entry_server.get()
                site_id = app.entry_site.get()
                if util.is_numeric(interface["site_id"]):
                    api_url = f"{server}/api/items/{interface['site_id']}/updatesite"
                else:
                    api_url = f"{server}/api/items/{site_id}/createsite"

                site_json["ApiVersion"] = 1.1
                site_json["ApiKey"] = api_key

            site_json["Title"] = interface["interface_name"]

            ### 詳細_編集要素 の 読み込みを開始 ###
            sht = workbook["詳細_編集要素"]
            app.log_output(f"#### 詳細_編集要素 の Sheet読み込みを開始")
            # 型情報データを取得する
            type_cells = [cell for cell in sht[config.EDIT_ROW_INDEX_TYPE] if not util.is_empty(cell.value)]
            type_cells.append(sht[util.get_address(config.EDIT_ROW_INDEX_TYPE, sht.max_column + 1)])
            column_name_list = []
            label_text_list = []
            for cur_cell, next_cell in util.pairwise(type_cells):
                app.log_output(f"{cur_cell} 読み込みを開始", "debug")

                start_row = config.EDIT_ROW_INDEX_GRID
                end_row = util.find_down_edge(sht, util.get_address(cur_cell.row, cur_cell.column)) + 1
                start_col = cur_cell.column
                end_col = next_cell.column - 1

                for i in range(start_row, end_row):
                    edit_obj = {}
                    for j in range(start_col + 1, end_col):

                        app.log_output(f"row: {i} col: {j} 読み込みを開始", "debug")
                        item_cell = sht.cell(row=config.EDIT_ROW_INDEX_ITEM, column=j)
                        target_cell = sht.cell(row=i, column=j)

                        # 値チェック
                        if util.is_empty(target_cell.value):
                            continue
                        elif item_cell.value not in config.PARAMETERS:
                            app.log_output(f"{util.get_address(item_cell.row, item_cell.column)}: PARAMETERS に登録されていないのでスキップします。: {item_cell.value}", "warning")
                            continue
                        # 型チェック
                        if config.PARAMETERS[item_cell.value]["type"] == "float":
                            if util.is_numeric(target_cell.value):
                                edit_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の型がfloatではないためスキップします。: {target_cell.value}", "warning")
                                continue
                        elif config.PARAMETERS[item_cell.value]["type"] == "bool":
                            if target_cell.value == config.MARK_OK or target_cell.value == "":
                                edit_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value == config.MARK_OK
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の型がboolではないためスキップします。: {target_cell.value}", "warning")
                                continue
                        elif config.PARAMETERS[item_cell.value]["type"] == "array":
                            if target_cell.value in config.PARAMETERS[item_cell.value]["values"]:
                                edit_obj[config.PARAMETERS[item_cell.value]["key"]] = config.PARAMETERS[item_cell.value]["values"][target_cell.value]
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の値が不正のためスキップします。: {target_cell.value}", "warning")
                                continue
                        else:
                            item_value = target_cell.value
                            if item_cell.value == "項目名":
                                if util.is_numeric(target_cell.value):
                                    item_value = config.TYPES[cur_cell.value]["key"] + util.to_N_digits(target_cell.value, 3)
                                elif len(target_cell.value) == 1 and target_cell.value.isalpha():
                                    item_value = config.TYPES[cur_cell.value]["key"] + target_cell.value.upper()
                                else:
                                    app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 不正な値なのでこの行はスキップします。: {target_cell.value}", "warning")
                                    break
                                column_name_list.append(item_value)
                            elif item_cell.value == "表示名":
                                label_text_list.append(item_value)

                            edit_obj[config.PARAMETERS[item_cell.value]["key"]] = item_value

                    if config.PARAMETERS["表示名"]["key"] in edit_obj:
                        # 必須チェック
                        try:
                            app.log_output(f'{edit_obj[config.PARAMETERS["項目名"]["key"]]} の項目名を取得しました。', "debug")
                        except NameError:
                            app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 項目名 が未入力なのでこの行はスキップします。", "warning")
                            continue
                        # 重複チェック
                        if util.is_duplicates(column_name_list):
                            name = column_name_list.pop()
                            app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 項目名 が重複しているのでこの行はスキップします。: {name}", "warning")
                            continue
                        if util.is_duplicates(label_text_list):
                            label = label_text_list.pop()
                            app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 表示名 が重複しているのでこの行はスキップします。: {label}", "warning")
                            continue

                        site_json["SiteSettings"]["Columns"].append(edit_obj)

            site_json["SiteSettings"]["EditorColumnHash"]["General"] = column_name_list

            ### 一覧_画面レイアウト の 読み込みを開始 ###
            for name in ["一覧_画面レイアウト", "詳細_画面レイアウト"]:
                sht = workbook[name]
                app.log_output(f"#### {name} の Sheet読み込みを開始")
                # 型情報データを取得する
                type_cells = [cell for cell in sht[config.LAYOUT_ROW_INDEX_TYPE] if not util.is_empty(cell.value)]
                type_cells.append(sht[util.get_address(config.LAYOUT_ROW_INDEX_TYPE, sht.max_column + 1)])
                for cur_cell, next_cell in util.pairwise(type_cells):
                    app.log_output(f"{cur_cell} 読み込みを開始", "debug")
                    start_row = config.LAYOUT_ROW_INDEX_GRID
                    end_row = util.find_down_edge(sht, util.get_address(cur_cell.row, cur_cell.column))
                    start_col = cur_cell.column
                    end_col = next_cell.column - 1
                    list = []
                    for i in range(start_row, end_row):
                        for j in range(start_col + 1, end_col):
                            app.log_output(f"row: {i} col: {j} 読み込みを開始", "debug")
                            target_cell = sht.cell(row=i, column=j)
                            if util.is_empty(target_cell.value):
                                continue
                            if target_cell.value not in label_text_list:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 詳細_編集要素 に登録されていないのでスキップします。: {target_cell.value} ", "warning")
                                continue
                            item_name = column_name_list[label_text_list.index(target_cell.value)]
                            if item_name in list:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: ２重に登録されているのでスキップします。: {target_cell.value} ", "warning")
                                continue
                            list.append(item_name)
                    if cur_cell.value in config.TABS:
                        if cur_cell.value == "編集要素":
                            site_json["SiteSettings"]["EditorColumnHash"]["General"] = list
                        elif cur_cell.value == "集計要素":
                            site_json["SiteSettings"]["Aggregations"] = list
                        else:
                            site_json["SiteSettings"][config.TABS[cur_cell.value]["key"]] = list


            ## スクリプト要素 の 読み込みを開始 ###
            sht = workbook["スクリプト要素"]
            app.log_output(f"#### スクリプト要素 の Sheet読み込みを開始")
            # 型情報データを取得する
            type_cells = [cell for cell in sht[config.EDIT_ROW_INDEX_TYPE] if not util.is_empty(cell.value)]
            type_cells.append(sht[util.get_address(config.EDIT_ROW_INDEX_TYPE, sht.max_column + 1)])
            for cur_cell, next_cell in util.pairwise(type_cells):

                app.log_output(f"{cur_cell} 読み込みを開始", "debug")
                start_row = config.EDIT_ROW_INDEX_GRID
                end_row = util.find_down_edge(sht, util.get_address(cur_cell.row, cur_cell.column)) + 1
                start_col = cur_cell.column
                end_col = next_cell.column - 1

                for i in range(start_row, end_row):
                    script_obj = {}
                    for j in range(start_col, end_col):
                        app.log_output(f"row: {i} col: {j} 読み込みを開始", "debug")
                        item_cell = sht.cell(row=config.EDIT_ROW_INDEX_ITEM, column=j)
                        target_cell = sht.cell(row=i, column=j)

                        # 値チェック
                        if util.is_empty(target_cell.value):
                            continue
                        elif item_cell.value not in config.PARAMETERS:
                            app.log_output(f"{util.get_address(item_cell.row, item_cell.column)}: PARAMETERS に登録されていないのでスキップします。: {item_cell.value}", "warning")
                            continue
                        # 型チェック
                        if config.PARAMETERS[item_cell.value]["type"] == "float":
                            if util.is_numeric(target_cell.value):
                                script_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の型がfloatではないためスキップします。: {target_cell.value}", "warning")
                                continue
                        elif config.PARAMETERS[item_cell.value]["type"] == "bool":
                            if target_cell.value == config.MARK_OK or target_cell.value == "":
                                script_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value == config.MARK_OK
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の型がboolではないためスキップします。: {target_cell.value}", "warning")
                                continue
                        elif config.PARAMETERS[item_cell.value]["type"] == "array":
                            if target_cell.value in config.PARAMETERS[item_cell.value]["values"]:
                                script_obj[config.PARAMETERS[item_cell.value]["key"]] = config.PARAMETERS[item_cell.value]["values"][target_cell.value]
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の値が不正のためスキップします。: {target_cell.value}", "warning")
                                continue
                        else:
                            script_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value


                    if config.PARAMETERS["タイトル"]["key"] in script_obj:
                        site_json["SiteSettings"][config.TABS[cur_cell.value]["key"]].append(script_obj)



            if app.radio_value.get() == "JSON":
                config.template_json["Sites"].append(site_json)

                # JSONデータを保存
                file_name = util.save_json(config.template_json, output_dir, interface["interface_name"])
                app.log_output(f"{file_name} を作成しました")
            elif app.radio_value.get() == "API":

                # JSONデータを用いてPOSTリクエストを実行
                response = requests.post(api_url, json=site_json)

                app.log_output(f'Status Code: {response.status_code}')
                app.log_output(f'Response: {response.json()}')

                app.log_output(f'{server}/items/{response.json()["Id"]}/index にアクセスしてください')



    except Exception as e:
        app.log_output(f"致命的なエラー: {e}", "error")

if __name__ == "__main__":
    app = gui.Gui("設計書からサイトパッケージ作成",execute_callback=main)
    app.mainloop()
