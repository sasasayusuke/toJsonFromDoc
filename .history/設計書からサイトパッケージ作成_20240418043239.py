import util
import gui
import config
import openpyxl
import os
import re
import requests
import copy


def main(app):

    try:

        # ファイル名のパターン（XXXX_テンプレート_画面設計書.xlsx）に合致するものを取得
        interfaces = []
        for file_name in os.listdir(config.DOCUMENT_DIRECTORY):
            match = re.match(r'.{4}_(.*)_画面設計書\.xlsx', file_name)
            if match:
                # パターンにマッチした場合
                interface_name = match.group(1)
                config.template_json["HeaderInfo"]["Convertors"][0]["SiteTitle"] = interface_name
                config.site_skeleton_json["Title"] = interface_name
                interfaces.append({
                    "interface_name": interface_name,
                    "file_name": file_name
                })

        for interface in interfaces:
            file_path = os.path.join(config.DOCUMENT_DIRECTORY, interface["file_name"])
            # ワークブックを読み込み
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # スケルトンのコピーを作成
            if app.radio_value.get() == "JSON":
                site_json = copy.deepcopy(config.site_skeleton_json)
            elif app.radio_value.get() == "API":
                site_json = copy.deepcopy(config.site_post_json)

            ### 詳細_編集要素 の 取り込み ###
            sht = workbook["詳細_編集要素"]
            app.log_output(f"詳細_編集要素 の 取り込みを開始")
            # 型情報データを取得する
            type_cells = [cell for cell in sht[config.EDIT_ROW_INDEX_TYPE] if not util.is_empty(cell.value)]
            type_cells.append(sht[util.get_address(config.EDIT_ROW_INDEX_TYPE, sht.max_column + 1)])
            column_name_list = []
            label_text_list = []
            for cur_cell, next_cell in util.pairwise(type_cells):
                start_row = config.EDIT_ROW_INDEX_GRID
                end_row = util.find_down_edge(sht, util.get_address(cur_cell.row, cur_cell.column))
                start_col = cur_cell.column
                end_col = next_cell.column - 1

                for i in range(start_row, end_row):
                    edit_obj = {}
                    for j in range(start_col + 1, end_col):
                        item_cell = sht.cell(row=config.EDIT_ROW_INDEX_ITEM, column=j)
                        target_cell = sht.cell(row=i, column=j)

                        # 値チェック
                        if util.is_empty(target_cell.value):
                            continue
                        elif item_cell.value not in config.PARAMETERS:
                            app.log_output(f"{util.get_address(item_cell.row, item_cell.column)}: PARAMETERS に登録されていないのでスキップします。: {item_cell.value}", "warning")
                            continue
                        elif item_cell.value == "項目名":
                            if isinstance(target_cell.value, (int, float)):
                                item_name = config.TYPES[cur_cell.value]["key"] + util.to_N_digits(target_cell.value, 3)
                            elif len(target_cell.value) == 1 and target_cell.value.isalpha():
                                item_name = config.TYPES[cur_cell.value]["key"] + target_cell.value.upper()
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 不正な値なのでこの行はスキップします。: {target_cell.value}", "warning")
                                break
                        # 型チェック
                        if config.PARAMETERS[item_cell.value]["type"] == "float":
                            if isinstance(target_cell.value, (int, float)):
                                edit_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の型がfloatではないためスキップします。: {target_cell.value}", "warning")
                                continue
                        elif config.PARAMETERS[item_cell.value]["type"] == "bool":
                            if target_cell.value == config.MARK_OK or target_cell.value == "":
                                edit_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value == config.MARK_OK
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の型がboolではないためスキップします。: {target_cell.value}", "warning")
                                continue
                        elif config.PARAMETERS[item_cell.value]["type"] == "array":
                            if target_cell.value in config.PARAMETERS[item_cell.value]["values"]:
                                edit_obj[config.PARAMETERS[item_cell.value]["key"]] = config.PARAMETERS[item_cell.value]["values"][target_cell.value]
                            else:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: {item_cell.value} の値が不正のためスキップします。: {target_cell.value}", "warning")
                                continue
                        else:
                            edit_obj[config.PARAMETERS[item_cell.value]["key"]] = target_cell.value

                    if config.PARAMETERS["表示名"]["key"] in edit_obj:
                        # 重複チェック
                        if item_name in column_name_list:
                            app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 項目名 が重複しているのでこの行はスキップします。: {item_name}", "warning")
                            continue
                        if edit_obj[config.PARAMETERS["表示名"]["key"]] in label_text_list:
                            app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 表示名 が重複しているのでこの行はスキップします。: {edit_obj[config.PARAMETERS["表示名"]["key"]]}", "warning")
                            continue

                        edit_obj[config.PARAMETERS["項目名"]["key"]] = item_name
                        column_name_list.append(item_name)
                        label_text_list.append(edit_obj[config.PARAMETERS["表示名"]["key"]])
                        site_json["SiteSettings"]["Columns"].append(edit_obj)

            site_json["SiteSettings"]["EditorColumnHash"]["General"] = column_name_list

            ### 一覧_画面レイアウト の 取り込み ###
            for name in ["一覧_画面レイアウト", "詳細_画面レイアウト"]:
                sht = workbook[name]
                app.log_output(f"{name} の 取り込みを開始")
                # 型情報データを取得する
                type_cells = [cell for cell in sht[config.LAYOUT_ROW_INDEX_TYPE] if not util.is_empty(cell.value)]
                type_cells.append(sht[util.get_address(config.LAYOUT_ROW_INDEX_TYPE, sht.max_column + 1)])
                for cur_cell, next_cell in util.pairwise(type_cells):
                    start_row = config.LAYOUT_ROW_INDEX_GRID
                    end_row = util.find_down_edge(sht, util.get_address(cur_cell.row, cur_cell.column))
                    start_col = cur_cell.column
                    end_col = next_cell.column - 1
                    list = []
                    for i in range(start_row, end_row):
                        edit_obj = {}
                        for j in range(start_col + 1, end_col):
                            target_cell = sht.cell(row=i, column=j)
                            if util.is_empty(target_cell.value):
                                continue
                            if target_cell.value not in label_text_list:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: 詳細_編集要素 に登録されていないのでスキップします。: {target_cell.value} ", "warning")
                                continue
                            item_name = column_name_list[label_text_list.index(target_cell.value)]
                            if item_name in list:
                                app.log_output(f"{util.get_address(target_cell.row, target_cell.column)}: ２重に登録されているのでスキップします。: {target_cell.value} ", "warning")
                                continue
                            list.append(item_name)
                    if cur_cell.value in config.TABS:
                        if cur_cell.value == "編集要素":
                            site_json["SiteSettings"]["EditorColumnHash"]["General"] = list
                        elif cur_cell.value == "集計要素":
                            site_json["SiteSettings"]["Aggregations"] = list
                            pass
                        else:
                            site_json["SiteSettings"][config.TABS[cur_cell.value]["key"]] = list



            if app.radio_value.get() == "JSON":
                config.template_json["Sites"].append(site_json)

                # JSONデータを保存
                file_name = util.save_json(config.template_json, config.DOCUMENT_DIRECTORY, interface["interface_name"])
                app.log_output(f"{file_name} を作成しました")
            elif app.radio_value.get() == "API":
                api_key = app.entry_api_key.get()
                server = app.entry_server.get()
                site_id = app.entry_site.get()
                url = f"{server}/api/items/{site_id}/createsite"

                site_json["ApiVersion"] = 1.1
                site_json["ApiKey"] = api_key

                # JSONデータを用いてPOSTリクエストを実行
                response = requests.post(url, json=site_json)

                app.log_output(f"{url} を作成しました")

                app.log_output(f'Status Code: {response.status_code}')
                app.log_output(f'Response: {response.json()}')



    except Exception as e:
        app.log_output(f"致命的なエラー: {e}", "error")

if __name__ == "__main__":
    app = gui.Gui("設計書からサイトパッケージ作成",execute_callback=main)
    app.mainloop()
